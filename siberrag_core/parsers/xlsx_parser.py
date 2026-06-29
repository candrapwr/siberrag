"""Parser untuk XLSX (.xlsx/.xlsm) berbasis openpyxl.

Setiap sheet menjadi satu TABLE; baris pertama diperlakukan sebagai header.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from siberrag_core.models.elements import Document, DocumentElement, ElementType
from siberrag_core.parsers.base import BaseParser, ParseError

try:
    from openpyxl import load_workbook  # type: ignore
    _HAS_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAS_OPENPYXL = False


class XlsxParser(BaseParser):
    extensions = ("xlsx", "xlsm")
    name = "xlsx"

    def parse(self, path: Path, *, filename: Optional[str] = None) -> Document:
        if not _HAS_OPENPYXL:
            raise ParseError("openpyxl tidak tersedia untuk parse XLSX.")
        try:
            wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        except Exception as exc:
            raise ParseError(f"Gagal parse XLSX {path.name}: {exc}") from exc

        root = DocumentElement.document(order=0)
        order = [0]

        for sheet in wb.worksheets:
            # heading = nama sheet
            root.add(DocumentElement.heading(sheet.title, level=1, page=1, order=order[0]))
            order[0] += 1
            table = DocumentElement(type=ElementType.TABLE, page_start=1, page_end=1, order=order[0])
            order[0] += 1
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if v is None else str(v).strip() for v in row]
                if all(c == "" for c in cells):
                    continue
                row_el = DocumentElement(type=ElementType.TABLE_ROW)
                for c in cells:
                    row_el.children.append(DocumentElement(type=ElementType.TABLE_CELL, content=c))
                table.children.append(row_el)
            if table.children:
                root.add(table)
        try:
            wb.close()
        except Exception:  # pragma: no cover
            pass
        return self._make_document(root, path=path, filename=filename)
